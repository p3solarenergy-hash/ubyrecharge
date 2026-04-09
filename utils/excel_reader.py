import os
import unicodedata
import warnings

import openpyxl
import pandas as pd
from openpyxl import load_workbook

from utils.drive_manifest import load_drive_manifest

warnings.filterwarnings("ignore")

APP_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
EXCEL_DIR = os.getenv("UBY_EXCEL_DIR", os.path.join(APP_DIR, "data"))
os.makedirs(EXCEL_DIR, exist_ok=True)

SKIP_FILES = {"desktop.ini"}
SKIP_PREFIX = ("~$",)
SKIP_DIRS = {"uby_recharge", "aurora_app", "__pycache__", ".claude"}


def ascii_key(s):
    """Normalize label to ASCII-safe lowercase for robust matching."""
    s = unicodedata.normalize("NFKD", str(s))
    s = s.encode("ascii", errors="ignore").decode().lower().strip()
    return s


def get_all_projects():
    """
    Scan the app data directory recursively and return all .xlsx project files.
    """
    files = []
    try:
        for root, dirnames, filenames in os.walk(EXCEL_DIR):
            dirnames[:] = [dirname for dirname in dirnames if dirname not in SKIP_DIRS]

            for filename in filenames:
                if (
                    filename.endswith(".xlsx")
                    and filename not in SKIP_FILES
                    and not any(filename.startswith(prefix) for prefix in SKIP_PREFIX)
                ):
                    full_path = os.path.join(root, filename)
                    relative_path = os.path.relpath(full_path, EXCEL_DIR)
                    files.append(relative_path)
    except OSError:
        pass
    return sorted(files)


def parse_inputs(filepath):
    """
    Parse inputs from project Excel (PLANILHA MODELO structure).
    Inputs sheet layout:
      col 0: label
      col 1: single/common value (for params common to all scenarios)
      col 2: pessimista value   (for scenario-specific params)
      col 3: conservador value  (for scenario-specific params)
      col 4: otimista value
    """
    inputs = {}
    try:
        wb = load_workbook(filepath, data_only=True)
        if "Inputs" not in wb.sheetnames:
            return inputs
        ws = wb["Inputs"]

        for row in ws.iter_rows(values_only=True):
            label_raw = row[0]
            if not label_raw or not str(label_raw).strip():
                continue
            label = str(label_raw).strip()
            akey = ascii_key(label)

            is_header = (
                label == label.upper() and len(label) > 8
                or any(header in akey for header in ["inputs centr", "parametro", "nan"])
            )
            if is_header:
                continue
            if label.startswith("â†’") or label.startswith("?") or label.startswith("AJUSTE"):
                continue

            val1 = row[1] if len(row) > 1 else None
            val_pess = row[2] if len(row) > 2 else None
            val_cons = row[3] if len(row) > 3 else None
            val_otim = row[4] if len(row) > 4 else None
            unit = ""
            if val_pess and isinstance(val_pess, str) and not _is_numeric(val_pess):
                unit = str(val_pess)
                val_pess = None

            if _is_numeric(val1):
                value = val1
                if not unit and _is_string(val_cons):
                    unit = str(val_cons) if val_cons else ""
            elif _is_numeric(val_cons):
                value = val_cons
            elif _is_numeric(val_pess):
                value = val_pess
            elif val1 and not isinstance(val1, (int, float)):
                value = val1
            else:
                continue

            for unit_col in [2, 5, 6]:
                cell = row[unit_col] if len(row) > unit_col else None
                if cell and isinstance(cell, str) and not _is_numeric(cell) and len(cell) < 30:
                    unit = cell
                    break

            inputs[label] = {
                "value": value,
                "unit": unit,
                "pessimista": val_pess if _is_numeric(val_pess) else value,
                "conservador": val_cons if _is_numeric(val_cons) else value,
                "otimista": val_otim if _is_numeric(val_otim) else value,
                "akey": akey,
            }
    except Exception:
        pass
    return inputs


def _is_numeric(v):
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    try:
        float(str(v))
        return True
    except (ValueError, TypeError):
        return False


def _is_string(v):
    return v is not None and isinstance(v, str) and not _is_numeric(v)


def parse_projection(filepath):
    try:
        wb = load_workbook(filepath, data_only=True)

        chart_sheet = None
        proj_sheet = None
        for sheet in wb.sheetnames:
            if "ChartData" in sheet or "_Chart" in sheet:
                chart_sheet = sheet
            if "Proje" in sheet and "Conservador" in sheet:
                proj_sheet = sheet

        target = chart_sheet or proj_sheet
        if not target:
            return None, None

        ws = wb[target]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return None, None

        headers = [str(header).strip() if header else f"col_{idx}" for idx, header in enumerate(rows[0])]
        data = [row for row in rows[1:] if any(cell is not None for cell in row)]
        df = pd.DataFrame(data, columns=headers)
        df = df.dropna(how="all")
        return df, target
    except Exception:
        return None, None


def parse_scenarios(filepath):
    try:
        wb = load_workbook(filepath, data_only=True)
        sheet_name = next((sheet for sheet in wb.sheetnames if "Cen" in sheet and "Conservador" in sheet), None)
        if not sheet_name:
            return None

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header_row = None
        for idx, row in enumerate(rows):
            if row[0] is not None and "Ocupa" in str(row[0]) and row[1] and "Horas" in str(row[1]):
                header_row = idx
                break
        if header_row is None:
            return None

        headers = [str(cell).strip() if cell else f"col_{idx}" for idx, cell in enumerate(rows[header_row])]
        data = [row for row in rows[header_row + 1 :] if row[0] is not None]
        df = pd.DataFrame(data, columns=headers).dropna(how="all")
        return df
    except Exception:
        return None


def parse_budget(filepath):
    try:
        wb = load_workbook(filepath, data_only=True)
        sheet_name = next((sheet for sheet in wb.sheetnames if "Orcamento" in sheet or "Or" in sheet and "amento" in sheet), None)
        if not sheet_name:
            return None, None

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header_row = None
        capex_total = None

        for idx, row in enumerate(rows):
            if row[0] is not None and str(row[0]).strip() == "#":
                header_row = idx
            if row[0] is not None and "CAPEX TOTAL" in str(row[0]).upper():
                for value in row:
                    if isinstance(value, (int, float)) and value > 0:
                        capex_total = value
                        break

        if header_row is None:
            return None, capex_total

        headers = [str(cell).strip() if cell else f"col_{idx}" for idx, cell in enumerate(rows[header_row])]
        data = [row for row in rows[header_row + 1 :] if row[0] is not None and str(row[0]).strip().isdigit()]
        if not data:
            return None, capex_total

        df = pd.DataFrame(data, columns=headers[: len(data[0])]).dropna(how="all")
        return df, capex_total
    except Exception:
        return None, None


def parse_full_project(filepath):
    name = os.path.splitext(os.path.basename(filepath))[0]
    inputs = parse_inputs(filepath)
    proj_df, proj_sheet = parse_projection(filepath)
    scen_df = parse_scenarios(filepath)
    budget_df, capex_total = parse_budget(filepath)

    kpis = {}
    for label, info in inputs.items():
        akey = info.get("akey", ascii_key(label))
        if "capex total" in akey and "repos" not in akey:
            kpis["capex"] = info["value"]
        elif "desconto" in akey or "tma" in akey:
            kpis["tma"] = info["value"]

    if capex_total:
        kpis["capex"] = capex_total

    return {
        "name": name,
        "filepath": filepath,
        "relative_path": os.path.relpath(filepath, EXCEL_DIR),
        "source": get_project_source(filepath),
        "inputs": inputs,
        "kpis": kpis,
        "projection": proj_df,
        "scenarios": scen_df,
        "budget": budget_df,
        "capex_total": capex_total,
    }


def get_project_source(filepath):
    relative_path = os.path.relpath(filepath, EXCEL_DIR)
    manifest = load_drive_manifest()
    return manifest.get(relative_path, {})


def save_inputs_to_excel(filepath, edited_inputs):
    """Write edited values back to Excel Inputs sheet."""
    try:
        wb = load_workbook(filepath)
        if "Inputs" not in wb.sheetnames:
            return False
        ws = wb["Inputs"]
        for row in ws.iter_rows():
            if not row[0].value:
                continue
            label = str(row[0].value).strip()
            if label in edited_inputs and len(row) > 1:
                col1_cell = row[1]
                col3_cell = row[3] if len(row) > 3 else None
                if col1_cell.value is not None and isinstance(col1_cell.value, (int, float)):
                    col1_cell.value = edited_inputs[label]
                elif col3_cell and col3_cell.value is not None and isinstance(col3_cell.value, (int, float)):
                    col3_cell.value = edited_inputs[label]
        wb.save(filepath)
        return True
    except Exception:
        return False
